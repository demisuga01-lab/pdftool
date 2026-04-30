import asyncio
import csv
import html
import mimetypes
import re
import zipfile
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status

from app.services.image_service import ImageService
from app.services.ocr_service import OCRService
from app.services.pdf_service import PDFService

PDF_FORMATS = {"pdf"}
OFFICE_FORMATS = {
    "doc",
    "docx",
    "xls",
    "xlsx",
    "ppt",
    "pptx",
    "odt",
    "ods",
    "odp",
    "rtf",
}
TEXT_FORMATS = {"txt", "html", "htm", "csv"}
IMAGE_FORMATS = {"jpg", "jpeg", "png", "webp", "avif", "tif", "tiff", "bmp"}
SVG_FORMATS = {"svg"}


class ConversionService:
    def _safe_filename(self, value: str | None, fallback_stem: str, extension: str) -> str:
        requested = Path(str(value or "").strip()).name
        safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(requested).stem if requested else fallback_stem).strip(".-") or fallback_stem
        return f"{safe_stem}.{extension.lstrip('.')}"

    def _normalize_format(self, value: str | None, *, fallback: str | None = None) -> str:
        normalized = (value or fallback or "").lower().strip().lstrip(".")
        # JPEG is the IANA name; "jpg" is the common file extension. Treat them
        # as the same canonical format internally.
        if normalized in {"jpeg", "jpg"}:
            return "jpg"
        if normalized in {"tif", "tiff"}:
            return "tiff"
        if normalized == "htm":
            return "html"
        if normalized in {"text", "plain"}:
            return "txt"
        return normalized

    def media_type_for(self, target_format: str) -> str:
        """Map a canonical output format to its IANA media type."""

        normalized = self._normalize_format(target_format)
        return {
            "jpg": "image/jpeg",
            "png": "image/png",
            "webp": "image/webp",
            "avif": "image/avif",
            "tiff": "image/tiff",
            "bmp": "image/bmp",
            "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "txt": "text/plain",
            "html": "text/html",
            "csv": "text/csv",
            "zip": "application/zip",
            "svg": "image/svg+xml",
            "eps": "application/postscript",
        }.get(normalized, "application/octet-stream")

    def detect_input_format(
        self,
        input_path: str | Path,
        from_format: str | None = None,
        mime_type: str | None = None,
    ) -> str:
        if from_format:
            return self._normalize_format(from_format)

        suffix = self._normalize_format(Path(input_path).suffix)
        if suffix:
            return suffix

        if mime_type:
            guessed = mime_type.split("/")[-1]
            return self._normalize_format(guessed)

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not detect the input format for conversion.",
        )

    def _safe_output_path(self, output_dir: Path, filename: str) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir / filename

    def _single_result(self, output_path: Path, *, original_name: str | None = None) -> dict[str, Any]:
        media_type = mimetypes.guess_type(output_path.name)[0] or "application/octet-stream"
        return {
            "output_path": str(output_path),
            "output_filename": output_path.name,
            "media_type": media_type,
            "original_name": original_name,
            "extension": output_path.suffix.lstrip("."),
        }

    def _multi_result(self, output_paths: list[Path], *, output_filename: str, original_name: str | None = None) -> dict[str, Any]:
        return {
            "output_paths": [str(path) for path in output_paths],
            "output_filename": output_filename,
            "media_type": "application/zip",
            "original_name": original_name,
            "extension": "zip",
        }

    async def _csv_to_xlsx(self, input_path: Path, output_path: Path) -> str:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        def write() -> None:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "CSV"
            with input_path.open("r", encoding="utf-8-sig", newline="") as source:
                reader = csv.reader(source)
                for row_index, row in enumerate(reader, start=1):
                    for column_index, cell in enumerate(row, start=1):
                        worksheet.cell(row=row_index, column=column_index, value=cell)
            workbook.save(output_path)

        await asyncio.to_thread(write)
        return str(output_path)

    async def _xlsx_to_csv(self, input_path: Path, output_path: Path) -> str:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        def write() -> None:
            from openpyxl import load_workbook

            workbook = load_workbook(input_path, read_only=True, data_only=True)
            worksheet = workbook.active
            with output_path.open("w", encoding="utf-8", newline="") as target:
                writer = csv.writer(target)
                for row in worksheet.iter_rows(values_only=True):
                    writer.writerow(["" if value is None else value for value in row])
            workbook.close()

        await asyncio.to_thread(write)
        return str(output_path)

    async def _text_to_html(self, input_path: Path, output_path: Path) -> str:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        text = await asyncio.to_thread(input_path.read_text, "utf-8", "ignore")
        body = html.escape(text)
        await asyncio.to_thread(
            output_path.write_text,
            f"<!doctype html><meta charset=\"utf-8\"><pre>{body}</pre>",
            "utf-8",
        )
        return str(output_path)

    async def _html_to_txt(self, input_path: Path, output_path: Path) -> str:
        import re

        output_path.parent.mkdir(parents=True, exist_ok=True)
        content = await asyncio.to_thread(input_path.read_text, "utf-8", "ignore")
        content = re.sub(r"<script\b[^>]*>.*?</script>", "", content, flags=re.IGNORECASE | re.DOTALL)
        content = re.sub(r"<style\b[^>]*>.*?</style>", "", content, flags=re.IGNORECASE | re.DOTALL)
        content = re.sub(r"<[^>]+>", " ", content)
        content = html.unescape(re.sub(r"\s+", " ", content)).strip()
        await asyncio.to_thread(output_path.write_text, content + "\n", "utf-8")
        return str(output_path)

    async def _zip_single(self, input_path: Path, output_path: Path) -> str:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        def write() -> None:
            with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
                archive.write(input_path, arcname=input_path.name)

        await asyncio.to_thread(write)
        return str(output_path)

    async def convert(
        self,
        input_path: str | Path,
        output_dir: str | Path,
        to_format: str,
        *,
        from_format: str | None = None,
        mime_type: str | None = None,
        settings: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        source = Path(input_path)
        destination_root = Path(output_dir)
        options = settings or {}
        input_format = self.detect_input_format(source, from_format, mime_type)
        target_format = self._normalize_format(to_format)
        requested_name = str(options.get("output_filename") or "").strip()
        original_name = source.name

        if input_format in PDF_FORMATS:
            if target_format == "docx":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "docx"))
                await PDFService().pdf_to_docx(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if target_format == "xlsx":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "xlsx"))
                await PDFService().pdf_to_excel(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if target_format == "txt":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "txt"))
                await PDFService().extract_text(source, output_path, layout=True, output_format="txt")
                return self._single_result(output_path, original_name=original_name)
            if target_format == "html":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "html"))
                await PDFService().pdf_to_html(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if target_format in {"png", "jpg", "webp"}:
                dpi = int(options.get("dpi", 180))
                jpeg_quality = int(options.get("quality", 85))
                page_base = re.sub(r"[^A-Za-z0-9._-]+", "-", (requested_name or source.stem)).strip(".-") or source.stem
                if target_format in {"png", "jpg"}:
                    outputs = [
                        Path(path)
                        for path in await PDFService().pdf_to_images(
                            source,
                            destination_root,
                            dpi=dpi,
                            format="jpeg" if target_format == "jpg" else target_format,
                            jpeg_quality=jpeg_quality,
                            transparent=bool(options.get("transparent", False)),
                        )
                    ]
                    zip_name = self._safe_filename(page_base, source.stem, "zip")
                    return self._multi_result(outputs, output_filename=zip_name, original_name=original_name)

                rendered_dir = destination_root / "rendered-png"
                rendered = [
                    Path(path)
                    for path in await PDFService().pdf_to_images(
                        source,
                        rendered_dir,
                        dpi=dpi,
                        format="png",
                        jpeg_quality=jpeg_quality,
                        transparent=bool(options.get("transparent", False)),
                    )
                ]
                converted: list[Path] = []
                for index, image_path in enumerate(rendered, start=1):
                    output_path = self._safe_output_path(destination_root, f"{page_base}-{index:04d}.webp")
                    await ImageService().convert_image(
                        image_path,
                        output_path,
                        "webp",
                        jpeg_quality,
                        bool(options.get("preserve_metadata", False)),
                        str(options.get("color_space", "srgb")),
                    )
                    converted.append(output_path)
                zip_name = self._safe_filename(page_base, source.stem, "zip")
                return self._multi_result(converted, output_filename=zip_name, original_name=original_name)
            if target_format in {"searchable_pdf", "pdf"}:
                output_name = self._safe_filename(requested_name, source.stem, "pdf")
                output_path = await OCRService().ocr(
                    source,
                    destination_root,
                    str(options.get("language", "eng")),
                    "pdf",
                    int(options.get("dpi", 300)),
                )
                final_path = Path(output_path)
                if final_path.name != output_name:
                    renamed = destination_root / output_name
                    final_path.replace(renamed)
                    final_path = renamed
                return self._single_result(final_path, original_name=original_name)

        if input_format in OFFICE_FORMATS or input_format in {"txt", "html", "rtf"}:
            if target_format == "pdf":
                output_path = Path(await PDFService().office_to_pdf(source, destination_root))
                desired = destination_root / self._safe_filename(requested_name, source.stem, "pdf")
                if output_path.name != desired.name:
                    output_path.replace(desired)
                    output_path = desired
                return self._single_result(output_path, original_name=original_name)
            if input_format == "xlsx" and target_format == "csv":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "csv"))
                await self._xlsx_to_csv(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if input_format in {"txt", "html"} and target_format == "zip":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "zip"))
                await self._zip_single(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if input_format == "txt" and target_format == "html":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "html"))
                await self._text_to_html(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if input_format == "html" and target_format == "txt":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "txt"))
                await self._html_to_txt(source, output_path)
                return self._single_result(output_path, original_name=original_name)

        if input_format == "csv":
            if target_format == "pdf":
                output_path = Path(await PDFService().office_to_pdf(source, destination_root))
                desired = destination_root / self._safe_filename(requested_name, source.stem, "pdf")
                if output_path.name != desired.name:
                    output_path.replace(desired)
                    output_path = desired
                return self._single_result(output_path, original_name=original_name)
            if target_format == "xlsx":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "xlsx"))
                await self._csv_to_xlsx(source, output_path)
                return self._single_result(output_path, original_name=original_name)
            if target_format == "zip":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "zip"))
                await self._zip_single(source, output_path)
                return self._single_result(output_path, original_name=original_name)

        if input_format in SVG_FORMATS:
            if target_format in {"png", "pdf", "eps"}:
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, target_format))
                await ImageService().convert_image(
                    source,
                    output_path,
                    target_format,
                    int(options.get("quality", 85)),
                    bool(options.get("preserve_metadata", False)),
                    str(options.get("color_space", "srgb")),
                )
                return self._single_result(output_path, original_name=original_name)

        if input_format in IMAGE_FORMATS:
            if target_format in {"jpg", "png", "webp", "avif", "tiff", "bmp"}:
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, target_format))
                converted = await ImageService().convert_image(
                    source,
                    output_path,
                    target_format,
                    int(options.get("quality", 85)),
                    bool(options.get("preserve_metadata", False)),
                    str(options.get("color_space", "srgb")),
                )
                if isinstance(converted, dict):
                    return converted
                return self._single_result(Path(converted), original_name=original_name)
            if target_format == "pdf":
                output_path = self._safe_output_path(destination_root, self._safe_filename(requested_name, source.stem, "pdf"))
                await PDFService().images_to_pdf([source], output_path)
                return self._single_result(output_path, original_name=original_name)

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Conversion from {input_format} to {target_format} is not supported yet.",
        )

    def select_queue(self, input_format: str, to_format: str) -> str:
        normalized_input = self._normalize_format(input_format)
        normalized_output = self._normalize_format(to_format)
        if normalized_output in {"docx", "xlsx", "searchable_pdf"}:
            return "heavy"
        if normalized_input in OFFICE_FORMATS or normalized_input in {"csv"}:
            return "heavy"
        return "fast"
