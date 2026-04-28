import asyncio
import csv
import json
import logging
import re
import shutil
from pathlib import Path
from typing import Literal

from fastapi import HTTPException, status

logger = logging.getLogger(__name__)

PdfQuality = Literal["screen", "ebook", "printer", "prepress"]
ImageFormat = Literal["png", "jpg", "jpeg", "pam", "pbm", "pgm", "ppm", "pnm"]
ColorMode = Literal["rgb", "cmyk", "gray"]
CompatibilityLevel = Literal["1.4", "1.5", "1.6", "1.7"]


class PDFService:
    async def _run_command(
        self,
        command: list[str],
        *,
        failure_status: int = status.HTTP_500_INTERNAL_SERVER_ERROR,
    ) -> None:
        logger.info("Running command: %s", " ".join(command))

        try:
            process = await asyncio.create_subprocess_exec(
                *command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await process.communicate()
        except FileNotFoundError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Required command not found: {command[0]}",
            ) from exc
        except OSError as exc:
            raise HTTPException(
                status_code=failure_status,
                detail=f"Failed to run command: {command[0]}",
            ) from exc

        if process.returncode != 0:
            error = stderr.decode("utf-8", errors="replace").strip()
            output = stdout.decode("utf-8", errors="replace").strip()
            detail = error or output or f"{command[0]} exited with code {process.returncode}"
            raise HTTPException(status_code=failure_status, detail=detail)

    async def _run_command_output(
        self,
        command: list[str],
        *,
        failure_status: int = status.HTTP_500_INTERNAL_SERVER_ERROR,
    ) -> tuple[str, str]:
        logger.info("Running command: %s", " ".join(command))

        try:
            process = await asyncio.create_subprocess_exec(
                *command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await process.communicate()
        except FileNotFoundError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Required command not found: {command[0]}",
            ) from exc
        except OSError as exc:
            raise HTTPException(
                status_code=failure_status,
                detail=f"Failed to run command: {command[0]}",
            ) from exc

        stdout_text = stdout.decode("utf-8", errors="replace")
        stderr_text = stderr.decode("utf-8", errors="replace")

        if process.returncode != 0:
            detail = stderr_text.strip() or stdout_text.strip() or f"{command[0]} exited with code {process.returncode}"
            raise HTTPException(status_code=failure_status, detail=detail)

        return stdout_text, stderr_text

    def _ensure_parent(self, output_path: str | Path) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    async def _strip_pdf_metadata(self, path: Path) -> Path:
        sanitized_path = path.with_name(f"{path.stem}-sanitized{path.suffix}")

        def write() -> None:
            from pypdf import PdfReader, PdfWriter

            reader = PdfReader(str(path))
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            writer.add_metadata({})
            with sanitized_path.open("wb") as output_file:
                writer.write(output_file)

        await asyncio.to_thread(write)
        sanitized_path.replace(path)
        return path

    async def _linearize_pdf(self, path: Path) -> Path:
        linearized_path = path.with_name(f"{path.stem}-linearized{path.suffix}")
        await self._run_command(["qpdf", "--linearize", str(path), str(linearized_path)])
        linearized_path.replace(path)
        return path

    async def _set_pdf_title(self, path: Path, title: str) -> Path:
        if not title.strip():
            return path

        titled_path = path.with_name(f"{path.stem}-title{path.suffix}")
        command = [
            "gs",
            "-dBATCH",
            "-dNOPAUSE",
            "-sDEVICE=pdfwrite",
            f"-sTitle={title}",
            f"-sOutputFile={titled_path}",
            str(path),
        ]
        await self._run_command(command)
        titled_path.replace(path)
        return path

    def _format_output_name(self, pattern: str, original_stem: str, index: int) -> str:
        return pattern.replace("{n}", str(index)).replace("{original}", original_stem)

    def _normalize_ocr_output_format(self, output_format: str) -> str:
        normalized = output_format.lower().strip()
        aliases = {"text": "txt", "searchable_pdf": "pdf"}
        normalized = aliases.get(normalized, normalized)

        if normalized not in {"txt", "pdf", "json", "docx", "hocr", "html"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="output_format must be one of: txt, pdf, json, docx, hocr, html",
            )

        return normalized

    def _parse_tsv_page(self, tsv_content: str, fallback_page_number: int) -> dict[str, object]:
        reader = csv.DictReader(tsv_content.splitlines(), delimiter="\t")
        words: list[dict[str, object]] = []

        for row in reader:
            text = (row.get("text") or "").strip()
            if row.get("level") != "5" or not text:
                continue

            words.append(
                {
                    "text": text,
                    "confidence": float(row.get("conf") or -1),
                    "left": int(row.get("left") or 0),
                    "top": int(row.get("top") or 0),
                    "width": int(row.get("width") or 0),
                    "height": int(row.get("height") or 0),
                    "block_num": int(row.get("block_num") or 0),
                    "par_num": int(row.get("par_num") or 0),
                    "line_num": int(row.get("line_num") or 0),
                    "word_num": int(row.get("word_num") or 0),
                }
            )

        return {
            "page": fallback_page_number,
            "words": words,
        }

    async def _extract_text_content(self, input_path: str | Path, layout: bool = True) -> str:
        command = ["pdftotext"]
        if layout:
            command.append("-layout")
        command.extend([str(input_path), "-"])
        stdout, _ = await self._run_command_output(command)
        return stdout

    async def _write_docx(self, text: str, output_path: str | Path) -> str:
        output = self._ensure_parent(output_path)

        def write() -> None:
            from docx import Document

            document = Document()
            paragraphs = text.splitlines() or [text]
            for paragraph in paragraphs:
                document.add_paragraph(paragraph)
            document.save(output)

        await asyncio.to_thread(write)
        return str(output)

    async def _write_json(self, payload: dict[str, object], output_path: str | Path) -> str:
        output = self._ensure_parent(output_path)
        await asyncio.to_thread(output.write_text, json.dumps(payload, indent=2), "utf-8")
        return str(output)

    def _split_page_tokens(self, pages: str) -> list[str]:
        tokens = [token.strip() for token in pages.split(",") if token.strip()]

        if not tokens:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one page selection is required",
            )

        for token in tokens:
            if not re.fullmatch(r"\d+(?:-\d+)?", token):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Page ranges must look like 1, 3-5, 8",
                )

        return tokens

    def _expand_page_tokens(self, tokens: list[str]) -> list[str]:
        pages: list[str] = []

        for token in tokens:
            if "-" not in token:
                pages.append(token)
                continue

            start_text, end_text = token.split("-", 1)
            start = int(start_text)
            end = int(end_text)
            if end < start:
                start, end = end, start
            pages.extend(str(page) for page in range(start, end + 1))

        return pages

    async def compress_pdf(
        self,
        input_path: str | Path,
        output_path: str | Path,
        quality: PdfQuality = "ebook",
        color_mode: ColorMode = "rgb",
        compatibility_level: CompatibilityLevel = "1.4",
        remove_metadata: bool = False,
        flatten_transparency: bool = False,
        linearize: bool = True,
        force_recompress: bool = False,
    ) -> dict[str, str | int | float]:
        if quality not in {"screen", "ebook", "printer", "prepress"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="quality must be one of: screen, ebook, printer, prepress",
            )
        if color_mode not in {"rgb", "cmyk", "gray"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="color_mode must be one of: rgb, cmyk, gray",
            )
        if compatibility_level not in {"1.4", "1.5", "1.6", "1.7"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="compatibility_level must be one of: 1.4, 1.5, 1.6, 1.7",
            )

        input_file = Path(input_path)
        output = self._ensure_parent(output_path)
        color_strategy = {
            "rgb": ("RGB", "DeviceRGB"),
            "cmyk": ("CMYK", "DeviceCMYK"),
            "gray": ("Gray", "DeviceGray"),
        }[color_mode]
        command = [
            "gs",
            "-dBATCH",
            "-dNOPAUSE",
            "-sDEVICE=pdfwrite",
            f"-dPDFSETTINGS=/{quality}",
            f"-dCompatibilityLevel={compatibility_level}",
            f"-sColorConversionStrategy={color_strategy[0]}",
            f"-sProcessColorModel={color_strategy[1]}",
            "-dColorImageDownsampleType=/Bicubic",
            "-dColorImageResolution=150",
            "-dGrayImageResolution=150",
            "-dMonoImageResolution=300",
            f"-sOutputFile={output}",
            str(input_file),
        ]
        if flatten_transparency:
            command.insert(-2, "-dNOTRANSPARENCY")
        await self._run_command(command)

        if remove_metadata:
            await self._strip_pdf_metadata(output)
        if linearize:
            await self._linearize_pdf(output)

        original_size = input_file.stat().st_size
        output_size = output.stat().st_size
        optimized = output_size < original_size
        message = None

        if output_size >= original_size and not force_recompress:
            shutil.copy2(input_file, output)
            output_size = output.stat().st_size
            optimized = False
            message = "Original PDF was already smaller; kept original to avoid increasing size."

        saved_bytes = max(original_size - output_size, 0)
        saved_percent = round((saved_bytes / original_size) * 100, 2) if original_size else 0.0

        return {
            "output_path": str(output),
            "optimized": optimized,
            "message": message,
            "original_size": original_size,
            "output_size": output_size,
            "saved_bytes": saved_bytes,
            "saved_percent": saved_percent,
            "original_size_bytes": original_size,
            "compressed_size_bytes": output_size,
            "reduction_percent": saved_percent,
        }

    async def merge_pdfs(
        self,
        input_paths: list[str | Path],
        output_path: str | Path,
        add_bookmarks: bool = True,
        metadata_title: str = "",
    ) -> str:
        if len(input_paths) < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least two PDFs are required to merge",
            )

        output = self._ensure_parent(output_path)
        if add_bookmarks:
            primary_input = str(input_paths[0])
            command = ["qpdf", primary_input, "--pages", *[str(path) for path in input_paths], "--", str(output)]
        else:
            command = ["qpdf", "--empty", "--pages", *[str(path) for path in input_paths], "--", str(output)]
        await self._run_command(command)
        await self._set_pdf_title(output, metadata_title)
        return str(output)

    async def split_pdf(
        self,
        input_path: str | Path,
        output_dir: str | Path,
        pages: str,
        naming_pattern: str = "page-{n}",
        output_format: str = "ranges",
    ) -> list[str]:
        output_directory = Path(output_dir)
        output_directory.mkdir(parents=True, exist_ok=True)
        input_stem = Path(input_path).stem
        tokens = self._split_page_tokens(pages)
        selections = tokens if output_format == "ranges" else self._expand_page_tokens(tokens)
        outputs: list[str] = []

        for index, selection in enumerate(selections, start=1):
            output_name = self._format_output_name(naming_pattern, input_stem, index)
            output = output_directory / f"{output_name}.pdf"
            command = ["qpdf", str(input_path), "--pages", ".", selection, "--", str(output)]
            await self._run_command(command)
            outputs.append(str(output))

        return outputs

    async def rotate_pdf(
        self,
        input_path: str | Path,
        output_path: str | Path,
        angle: int,
        pages: str = "all",
    ) -> str:
        if angle not in {0, 90, 180, 270}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="angle must be one of: 0, 90, 180, 270",
            )

        output = self._ensure_parent(output_path)
        command = ["qpdf", str(input_path), f"--rotate={angle}:{pages}", str(output)]
        await self._run_command(command)
        return str(output)

    async def extract_text(
        self,
        input_path: str | Path,
        output_path: str | Path,
        layout: bool = True,
        output_format: str = "txt",
    ) -> str:
        normalized_output = self._normalize_ocr_output_format(output_format)
        output = self._ensure_parent(output_path)

        if normalized_output == "txt":
            command = ["pdftotext"]
            if layout:
                command.append("-layout")
            command.extend([str(input_path), str(output)])
            await self._run_command(command)
            return str(output)

        if normalized_output == "html":
            command = ["pdftotext", "-htmlmeta", str(input_path), str(output)]
            await self._run_command(command)
            return str(output)

        text = await self._extract_text_content(input_path, layout=layout)

        if normalized_output == "json":
            pages = [page.strip() for page in text.split("\f")]
            payload = {
                "text": text,
                "pages": [
                    {"page": index, "text": page}
                    for index, page in enumerate(pages, start=1)
                    if page
                ],
            }
            return await self._write_json(payload, output)

        return str(output)

    async def pdf_to_docx(
        self,
        input_path: str | Path,
        output_path: str | Path,
        layout: bool = True,
    ) -> str:
        text = await self._extract_text_content(input_path, layout=layout)
        return await self._write_docx(text, output_path)

    async def pdf_to_excel(
        self,
        input_path: str | Path,
        output_path: str | Path,
        layout: bool = True,
    ) -> str:
        output = self._ensure_parent(output_path)
        text = await self._extract_text_content(input_path, layout=layout)
        pages = [page for page in text.split("\f") if page.strip()]

        def write() -> None:
            from openpyxl import Workbook

            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            for page_number, page_text in enumerate(pages or [text], start=1):
                worksheet = workbook.create_sheet(title=f"Page {page_number}"[:31])
                for row_index, line in enumerate(page_text.splitlines(), start=1):
                    cells = [cell for cell in re.split(r"\s{2,}", line.strip()) if cell] or [line.strip()]
                    for column_index, cell in enumerate(cells, start=1):
                        worksheet.cell(row=row_index, column=column_index, value=cell)

            workbook.save(output)

        await asyncio.to_thread(write)
        return str(output)

    async def pdf_to_html(self, input_path: str | Path, output_path: str | Path) -> str:
        return await self.extract_text(input_path, output_path, layout=True, output_format="html")

    async def pdf_to_images(
        self,
        input_path: str | Path,
        output_dir: str | Path,
        dpi: int = 150,
        format: ImageFormat = "png",
        jpeg_quality: int = 82,
        transparent: bool = False,
    ) -> list[str]:
        if dpi < 1 or dpi > 1200:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="dpi must be between 1 and 1200",
            )

        output_directory = Path(output_dir)
        output_directory.mkdir(parents=True, exist_ok=True)
        output_pattern = output_directory / f"page-%03d.{format}"
        command = ["mutool", "draw", "-r", str(dpi), "-F", format, "-o", str(output_pattern)]
        if transparent and format == "png":
            command.extend(["-c", "rgba"])
        command.append(str(input_path))
        await self._run_command(command)
        outputs = sorted(output_directory.glob(f"page-*.{format}"))

        if format == "jpeg":
            import pyvips

            for path in outputs:
                image = pyvips.Image.new_from_file(str(path), access="sequential")
                image.write_to_file(str(path), Q=jpeg_quality)

        return [str(path) for path in outputs]

    async def images_to_pdf(self, input_paths: list[str | Path], output_path: str | Path) -> str:
        if not input_paths:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one image is required",
            )

        output = self._ensure_parent(output_path)
        logger.info("Running img2pdf conversion for %s image(s)", len(input_paths))

        try:
            import img2pdf

            def convert() -> None:
                with output.open("wb") as output_file:
                    output_file.write(img2pdf.convert([str(path) for path in input_paths]))

            await asyncio.to_thread(convert)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to convert images to PDF: {exc}",
            ) from exc

        return str(output)

    async def office_to_pdf(self, input_path: str | Path, output_dir: str | Path) -> str:
        output_directory = Path(output_dir)
        output_directory.mkdir(parents=True, exist_ok=True)
        command = [
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_directory),
            str(input_path),
        ]
        await self._run_command(command)
        return str(output_directory / f"{Path(input_path).stem}.pdf")

    async def encrypt_pdf(
        self,
        input_path: str | Path,
        output_path: str | Path,
        user_password: str,
        owner_password: str = "",
    ) -> str:
        if not user_password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="user_password is required",
            )

        output = self._ensure_parent(output_path)
        command = [
            "qpdf",
            "--encrypt",
            user_password,
            owner_password,
            "256",
            "--",
            str(input_path),
            str(output),
        ]
        await self._run_command(command)
        return str(output)

    async def decrypt_pdf(
        self,
        input_path: str | Path,
        output_path: str | Path,
        password: str,
    ) -> str:
        if not password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="password is required",
            )

        output = self._ensure_parent(output_path)
        command = ["qpdf", "--decrypt", f"--password={password}", str(input_path), str(output)]
        await self._run_command(command)
        return str(output)

    async def repair_pdf(self, input_path: str | Path, output_path: str | Path) -> str:
        output = self._ensure_parent(output_path)
        command = ["qpdf", str(input_path), str(output)]
        await self._run_command(command)
        return str(output)

    async def ocr_pdf(
        self,
        input_path: str | Path,
        output_dir: str | Path,
        language: str = "eng",
        output_format: str = "txt",
    ) -> str:
        logger.info("Starting ocr_pdf for file %s", input_path)
        normalized_output = self._normalize_ocr_output_format(output_format)
        output_directory = Path(output_dir)
        output_directory.mkdir(parents=True, exist_ok=True)
        page_pattern = output_directory / "page-%04d.png"

        await self._run_command(
            [
                "mutool",
                "draw",
                "-F",
                "png",
                "-r",
                "300",
                "-o",
                str(page_pattern),
                str(input_path),
            ]
        )

        page_images = sorted(output_directory.glob("page-*.png"))
        if not page_images:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to render PDF pages for OCR",
            )

        if normalized_output in {"txt", "docx"}:
            page_texts: list[str] = []
            for page_image in page_images:
                stdout, _ = await self._run_command_output(
                    ["tesseract", str(page_image), "stdout", "-l", language]
                )
                page_texts.append(stdout.strip())

            combined_text = "\n\n".join(text for text in page_texts if text).strip()
            if normalized_output == "docx":
                return await self._write_docx(combined_text, output_directory / "ocr-output.docx")
            output = self._ensure_parent(output_directory / "ocr-output.txt")
            await asyncio.to_thread(output.write_text, combined_text, "utf-8")
            return str(output)

        if normalized_output == "json":
            pages: list[dict[str, object]] = []
            for page_number, page_image in enumerate(page_images, start=1):
                output_base = output_directory / f"ocr-page-{page_number:04d}"
                await self._run_command(
                    ["tesseract", str(page_image), str(output_base), "-l", language, "tsv"]
                )
                tsv_content = output_base.with_suffix(".tsv").read_text(encoding="utf-8")
                pages.append(self._parse_tsv_page(tsv_content, page_number))

            return await self._write_json(
                {"language": language, "pages": pages},
                output_directory / "ocr-output.json",
            )

        if normalized_output == "hocr":
            output_files: list[Path] = []
            for page_number, page_image in enumerate(page_images, start=1):
                output_base = output_directory / f"ocr-page-{page_number:04d}"
                await self._run_command(
                    ["tesseract", str(page_image), str(output_base), "-l", language, "hocr"]
                )
                output_files.append(output_base.with_suffix(".hocr"))

            merged_output = self._ensure_parent(output_directory / "ocr-output.hocr")
            content = "".join(path.read_text(encoding="utf-8") for path in output_files)
            await asyncio.to_thread(merged_output.write_text, content, "utf-8")
            return str(merged_output)

        page_pdfs: list[Path] = []
        for page_number, page_image in enumerate(page_images, start=1):
            output_base = output_directory / f"ocr-page-{page_number:04d}"
            await self._run_command(
                ["tesseract", str(page_image), str(output_base), "-l", language, "pdf"]
            )
            page_pdfs.append(output_base.with_suffix(".pdf"))

        merged_output = self._ensure_parent(output_directory / "ocr-searchable.pdf")
        await self._run_command(
            ["qpdf", "--empty", "--pages", *[str(path) for path in page_pdfs], "--", str(merged_output)]
        )
        return str(merged_output)
